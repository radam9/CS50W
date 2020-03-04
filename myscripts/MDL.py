import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Variables
myinfo = "myscripts/myinfo.xlsx"
dbfile = "myscripts/dbtest.xlsx"

# Functions
def checkurls(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
    except:
        return f"The url ## {url} ## did not work!"
    return f"{url} ## is active!"


def getdramainfo(url):
    r = requests.get(url)
    data = []
    soup = BeautifulSoup(r.text, "html.parser")
    x = soup.find("ul", class_="list m-b-0")
    data.append(x.find("span", itemprop="name").contents[0])
    temp = x.find(text="Country:")
    data.append(temp.find_parent("li").text[9:-1])
    temp = x.find(text="Episodes:")
    data.append(temp.find_parent("li").text[10::])
    temp = x.find(text="Aired:")
    year = temp.find_parent("li").text
    data.append(year[year.find(",") + 2 : year.find(",") + 6])
    temp = x.find(text="Original Network:")
    if temp == None:
        data.append(None)
    else:
        data.append(temp.find_next().text)
    temp = x.find(text="Duration:")
    temp = temp.find_parent("li").text[10::]
    if len(temp) == 12:
        data.append(int(temp[0:1]) * 60 + int(temp[6:7]))
    elif len(temp) == 13:
        data.append(int(temp[0:1]) * 60 + int(temp[6:8]))
    else:
        data.append(int(temp[0:2]))
    return data


def appendinfofile(data):
    wb = load_workbook(dbfile)
    sheet = wb["test"]
    r = sheet.max_row + 1
    for i in range(10):
        sheet.cell(row=r, column=i + 1).value = data[i]

    wb.save(dbfile)


def populate(src):
    wb = load_workbook(src)
    sheet_drama = wb["Dramas"]
    for row in sheet_drama.iter_rows():
        data = []
        for cell in row:
            data.append(cell.value)
        newdata = getdramainfo(data[1])
        data.extend(newdata)
        appendinfofile(data)


populate(myinfo)
