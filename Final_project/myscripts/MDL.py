import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# setup django
import os, sys

sys.path.append("D:\Knowledge\CS50W\Projects\FinalProject")
os.environ["DJANGO_SETTINGS_MODULE"] = "dramas.settings"
import django

django.setup()
from mydramas.models import Network, Drama

# Variables
myinfo = "myscripts/myinfo.xlsx"
# dbfile = "myscripts/dbtest.xlsx"
networks = [
    "tvN",
    "SBS",
    "MBC",
    "KBS2",
    "jTBC",
    "OCN",
    "MBN",
    "On Style",
    "Naver TV Cast",
    "DramaX",
    "TV Chosun",
    "Line TV",
    "None",
]

# Functions
def checkurls(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
    except:
        return f"The url ## {url} ## did not work!"
    return f"{url} ## is active!"


def getdramainfo(url):
    r = requests.get(url)
    data = []
    soup = BeautifulSoup(r.text, "html.parser")
    x = soup.find("ul", class_="list m-b-0")
    data.append(x.find("span", itemprop="name").contents[0])
    # temp = x.find(text="Country:")
    # data.append(temp.find_parent("li").text[9:-1])
    temp = x.find(text="Episodes:")
    data.append(temp.find_parent("li").text[10::])
    temp = x.find(text="Aired:")
    year = temp.find_parent("li").text
    data.append(year[year.find(",") + 2 : year.find(",") + 6])
    temp = x.find(text="Original Network:")
    if temp == None:
        data.append(None)
    else:
        data.append(temp.find_next().text)
    temp = x.find(text="Duration:")
    temp = temp.find_parent("li").text[10::]
    if len(temp) == 12:
        data.append(int(temp[0:1]) * 60 + int(temp[6:7]))
    elif len(temp) == 13:
        data.append(int(temp[0:1]) * 60 + int(temp[6:8]))
    else:
        data.append(int(temp[0:2]))
    return data


def getdramainfoview(url):
    r = requests.get(url)
    data = []
    soup = BeautifulSoup(r.text, "html.parser")
    x = soup.find("ul", class_="list m-b-0")
    y = soup.find("div", class_="col-sm-4 film-cover cover")
    data.append(x.find("span", itemprop="name").contents[0])
    # temp = x.find(text="Country:")
    # data.append(temp.find_parent("li").text[9:-1])
    temp = x.find(text="Episodes:")
    data.append(temp.find_parent("li").text[10::])
    temp = x.find(text="Aired:")
    year = temp.find_parent("li").text
    data.append(year[year.find(",") + 2 : year.find(",") + 6])
    temp = x.find(text="Original Network:")
    if temp == None:
        data.append(None)
    else:
        data.append(temp.find_next().text)
    temp = x.find(text="Duration:")
    temp = temp.find_parent("li").text[10::]
    if len(temp) == 12:
        data.append(int(temp[0:1]) * 60 + int(temp[6:7]))
    elif len(temp) == 13:
        data.append(int(temp[0:1]) * 60 + int(temp[6:8]))
    else:
        data.append(int(temp[0:2]))
    image = y.find("img")
    data.append(image["src"])
    return data


def appendinfofile(data):
    wb = load_workbook(dbfile)
    sheet = wb["test"]
    r = sheet.max_row + 1
    for i in range(10):
        sheet.cell(row=r, column=i + 1).value = data[i]

    wb.save(dbfile)


def populate(src):
    wb = load_workbook(src)
    sheet_drama = wb["Dramas"]
    for row in sheet_drama.iter_rows():
        data = []
        for cell in row:
            data.append(cell.value)
        newdata = getdramainfo(data[1])
        data.extend(newdata)
        appendinfofile(data)


def populatedrama(src):
    wb = load_workbook(src)
    sheet_drama = wb["Dramas"]
    payload = []
    for row in sheet_drama.iter_rows():
        data = []
        for cell in row:
            data.append(cell.value)
        newdata = getdramainfo(data[1])
        data.extend(newdata)
        if data[2] == "Yes":
            data[2] = True
        else:
            data[2] = False
        info = {
            "rating": round(float(data[0]), 1),
            "mdlurl": data[1],
            "favorite": data[2],
            "watchdate": data[3].date(),
            "title": data[4],
            "epcount": int(data[5]),
            "year": int(data[6]),
            "network": Network.objects.get(title=str(data[7])),
            "eplength": int(data[8]),
        }
        # payload.append(Drama(**info))
        d = Drama(**info)
        d.save()
    # Drama.objects.bulk_create(payload)


# rating - url - favorite - date - title - epcount - year - network - eplength


def populatenetwork(network_name):
    t = Network.objects.filter(title=network_name).count()
    if t == 0:
        n = Network(title=network_name)
        n.save()


##Code used to populate the database from an excel file
# for i in networks:
#     populatenetwork(i)
# populatedrama(myinfo)
